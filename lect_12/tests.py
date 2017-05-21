from lect_12_1 import *
from openpyxl import load_workbook
import unittest


class FixturesTest(unittest.TestCase):
    def setUp(self):
        self.event_loop = asyncio.new_event_loop()
        self.url = URL
        asyncio.set_event_loop(self.event_loop)
        self.session = create_session()
        self.result_of_session = self.event_loop.run_until_complete(get_url(self.session, self.event_loop, self.url))

        self.table_head, self.table_body = search_of_data(self.result_of_session)
        self.headers = {
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8",
            "Accept-Encoding": "gzip, deflate",
            "Accept-Language": "ru-RU,ru;q=0.8,en-US;q=0.6,en;q=0.4",
            "Host": "httpbin.org",
            "Upgrade-Insecure-Requests": "1",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) "
                          "Chrome/57.0.2987.98 Safari/537.36 OPR/44.0.2510.857"
        }

    def test_create_session(self):
        self.assertEqual(self.session, aiohttp.ClientSession, 'create_session should return aiohttp client session '
                                                              'object, but return {}'.format(type(self.session)))

    def test_run_loop(self):
        self.assertIsInstance(self.result_of_session, str, 'Result of loop should return string, but '
                                                           'return {}'.format(type(self.result_of_session)))

    def test_search_of_data(self):
        self.assertIsInstance(self.table_head, tuple, 'Table head must be list, '
                                                      'but return type: {}'.format(type(self.table_head)))
        self.assertIsInstance(self.table_body, list, 'Table head must be list, '
                                                     'but return type: {}'.format(type(self.table_body)))
        self.assertGreater(len(self.table_head), 0, 'Table head list is empty')
        self.assertGreater(len(self.table_body), 0, 'Table body list is empty')

    def test_write_to_xls(self):
        wb = load_workbook('coin_market_cap.xlsx', read_only=True)
        ws = wb.active
        r = ws['A1':'J1']
        l_head = []
        l_first_row = []

        for row in r:
            for i in row:
                l_head.append(i.value)
        self.assertLessEqual(l_head, list(self.table_head), 'Table head was write wrong to xlsx file. '
                                                            'Head look like: {}'.format(list(self.table_head)))
        # r = ws['A2':'J2']
        # for row in r:
        #     for i in row:
        #         l_first_row.append(i.value)
        # self.assertLessEqual(l_first_row, self.table_body[0], 'Table first row was write wrong to xlsx file. '
        #                                                             'First row looks like: {}'.format(self.table_body[0]))
        wb.close()


if __name__ == '__main__':
    unittest.main()
